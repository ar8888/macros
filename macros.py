import openpyxl
import logs
import pandas


class Macros:
    file_in = ''
    file_in_name = ''
    type_out = ''
    list_manuf = []
    num__header = 0
    col_name = 0
    data = None
    dict = {}
    message_box = None

    def __init__(self, msqln):
        self.message_box = msqln

    def set_options(self, params):
        self.file_in = params['file_in']
        self.file_in_name = params['file_in_name']
        self.type_out = params['type_out']
        self.list_manuf = params['list_manuf']
        self.num_header = params['num_header'] - 1
        self.col_name = params['col_name'] - 1
        self.folder_out = params['folder_out']
        if params['col_fn'] != 0:
            self.col_fn = params['col_fn'] - 1
        else:
            self.col_fn = 0

    def run(self):
        self.load_dict()
        self.load_file()
        if (self.col_name in self.data.columns) is False:
            self.message_box.emit(f'Не найдена колонка {self.col_name} в загруженном файле')
            return False
        if self.type_out == "one":
            self.save_one_file()
        else:
            self.save_many_files()
        self.data = None
        return True

    def load_dict(self):
        f_conv = "dict.xlsx"
        wb = openpyxl.load_workbook(f_conv)
        sheet = wb.active
        max_row = sheet.max_row
        for i in range(1, max_row):
            val1 = '' if sheet.cell(row=i, column=1).value is None else sheet.cell(row=i, column=1).value
            val3 = '' if sheet.cell(row=i, column=3).value is None else sheet.cell(row=i, column=3).value
            manuf = str(val1).replace("'", "").strip()
            search_txt = str(val3).replace("'", "").strip().casefold()
            if search_txt == '' or manuf == '':
                continue
            if manuf in self.dict:
                self.dict[manuf].add(search_txt)
            else:
                self.dict[manuf] = set()
                self.dict[manuf].add(search_txt)

    def load_file(self):
        self.data = pandas.read_excel(self.file_in, skiprows=self.num_header)
        cols = self.data.columns
        col_for_search = cols[self.col_name].upper()
        self.message_box.emit(f'Поиск осуществляем по колонке {col_for_search}')
        max_row = len(self.data.index)
        size_matrix = self.data.shape
        ml = ('') * max_row
        self.data.insert(size_matrix[1], "manuf", ml)
        self.col_name = self.data.columns.values.tolist()[self.col_name]
        if self.col_fn != 0:
            print('меняем')
            self.data[cols[self.col_fn]] = self.data[cols[self.col_fn]].astype('str')

    def save_many_files(self):
        max_row = len(self.data.index)
        for manuf_name in self.list_manuf:
            if manuf_name in self.dict is False:
                logs.write_log(f'не найден производитель {manuf_name} в справочнике')
                self.message_box.emit(f'не найден производитель {manuf_name} в справочнике')
                continue
            self.message_box.emit(f"формируем данные по производителю {manuf_name}")
            for i in range(max_row):
                name_good = str(self.data.iloc[i][self.col_name]).casefold()
                for search_txt in self.dict[manuf_name]:
                    if name_good.find(search_txt) != -1:
                        self.data.at[i, 'manuf'] = manuf_name
                        break
            # записываем данные в файл
            select_ = self.data[self.data.manuf == manuf_name]
            select_ = select_.drop(columns='manuf')
            writer = pandas.ExcelWriter(f"{self.folder_out}/{self.file_in_name} & {manuf_name}.xlsx",
                                        engine='xlsxwriter', date_format=r'dd.mm.yyyy',
                                        datetime_format=r'dd.mm.yyyy HH:MM:SS')
            select_.to_excel(writer, index=False)
            writer.close()

    def save_one_file(self):
        list_search = set()
        list_manuf = []
        max_row = len(self.data.index)
        for manuf_name in self.list_manuf:
            if manuf_name in self.dict is False:
                self.message_box.emit(f'не найден производитель {manuf_name} в справочнике')
                logs.write_log(f'не найден производитель {manuf_name} в справочнике')
                continue
            list_manuf.append(manuf_name)
            for search_txt in self.dict[manuf_name]:
                list_search.add(search_txt)
        for i in range(max_row):
            name_good = str(self.data.iloc[i][self.col_name]).casefold()
            for search_txt in list_search:
                if name_good.find(search_txt) != -1:
                    self.data.at[i, 'manuf'] = 'Отбор'
                    break
        # записываем данные в файл
        # name_file = " ".join(list_manuf)
        file_result = f"{self.folder_out}/{self.file_in_name} & результат.xlsx"
        writer = pandas.ExcelWriter(file_result, engine='xlsxwriter', date_format=r'dd.mm.yyyy',
                                    datetime_format=r'dd.mm.yyyy HH:MM:SS')
        select_ = self.data[self.data.manuf == 'Отбор']
        select_ = select_.drop(columns='manuf')
        select_.to_excel(writer, 'Данные', index=False)
        # дописываем на оттельный лист производителей
        proizv = pandas.DataFrame(list_manuf)
        proizv.to_excel(writer, 'Производители', header=['Производители'], startrow=5, index=False)
        writer.close()


if __name__ == "__main__":
    worker = Macros()
    worker.set_options("бюджетка.xlsx", "many", "result", ["Санофи", "Ксантис"], 1, 10)
    worker.run()
